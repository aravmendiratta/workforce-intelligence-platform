import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference

file_path = 'data/processed/master_health_data.xlsx'
print(f"Loading {file_path} to add formulas and static pivot tables...")

# 1. Load data with Pandas to create static Pivot Tables
df = pd.read_excel(file_path)
df['Calculated_ABI_Score'] = (df['CurrentWorkAbility'] * 2) + (df['WorkAbilityPhysical'] * 3) + (df['WorkAbilityMental'] * 2) + 4
df['AgeGroup'] = pd.cut(df['Age'], bins=range(20, 80, 10), right=False, labels=[f"{i}-{i+9}" for i in range(20, 70, 10)])

# Pivot 1: Average ABI by Department and Age Group
pivot1 = df.pivot_table(values='Calculated_ABI_Score', index='Department', columns='AgeGroup', aggfunc='mean').round(1)

# Pivot 2: Correlation between Sick Days and Stress Levels
pivot2 = df.pivot_table(values='SickDaysPastYear', index='StressLevel', aggfunc='mean').round(1)

# 2. Modify Excel using openpyxl
wb = load_workbook(file_path)
ws_data = wb.active
ws_data.title = "Master Data"

# Add formula columns headers
max_col = ws_data.max_column
ws_data.cell(row=1, column=max_col + 1, value='Calculated_ABI_Score')
ws_data.cell(row=1, column=max_col + 2, value='ABI_Category')

# Get column letters
curr_work_col = 'K' # CurrentWorkAbility
phys_work_col = 'L' # WorkAbilityPhysical
ment_work_col = 'M' # WorkAbilityMental

# Add Excel formulas to all rows
print("Adding Excel formulas...")
for row in range(2, ws_data.max_row + 1):
    # Calculated ABI Score formula
    formula_score = f"=({curr_work_col}{row} * 2) + ({phys_work_col}{row} * 3) + ({ment_work_col}{row} * 2) + 4"
    ws_data.cell(row=row, column=max_col + 1, value=formula_score)
    
    # IFS formula for Category (Changed to nested IF for compatibility with all Excel versions)
    l_col = ws_data.cell(row=row, column=max_col + 1).coordinate
    formula_ifs = f'=IF({l_col}<28, "Poor", IF({l_col}<37, "Moderate", IF({l_col}<44, "Good", "Excellent")))'
    ws_data.cell(row=row, column=max_col + 2, value=formula_ifs)

# Add Pivot Tables sheet
print("Adding Pivot Tables and Charts...")
ws_pivot = wb.create_sheet(title="Pivot Tables & Charts")

# Write Pivot 1
ws_pivot.cell(row=1, column=1, value="Avg ABI by Dept and Age")
for r in dataframe_to_rows(pivot1, index=True, header=True):
    ws_pivot.append(r)

# Write Pivot 2
start_row_p2 = len(pivot1) + 5
ws_pivot.cell(row=start_row_p2, column=1, value="Avg Sick Days by Stress Level")
ws_pivot.cell(row=start_row_p2+1, column=1, value="Stress Level")
ws_pivot.cell(row=start_row_p2+1, column=2, value="Avg Sick Days")

for r_idx, (index, row) in enumerate(pivot2.iterrows(), start=start_row_p2+2):
    ws_pivot.cell(row=r_idx, column=1, value=index)
    ws_pivot.cell(row=r_idx, column=2, value=row['SickDaysPastYear'])

# Create Combo Chart (Line Chart for this one)
chart = LineChart()
chart.title = "Impact of Stress on Sick Days"
chart.style = 13
chart.y_axis.title = "Avg Sick Days"
chart.x_axis.title = "Stress Level"

# Data for chart
data = Reference(ws_pivot, min_col=2, min_row=start_row_p2+1, max_row=start_row_p2+1+len(pivot2))
cats = Reference(ws_pivot, min_col=1, min_row=start_row_p2+2, max_row=start_row_p2+1+len(pivot2))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws_pivot.add_chart(chart, "E5")

wb.save('data/processed/final_health_dashboard_data.xlsx')
print("Saved data/processed/final_health_dashboard_data.xlsx successfully!")
